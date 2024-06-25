import openpyxl

excel_dataframe = openpyxl.load_workbook("../data/notasPrueba.xlsx")

dataframe = excel_dataframe.active


# esto imprime el nombre de la hoja, esto sirve apra comprobar la hoja con la que estamos trabajando y verificar si accede correctamente al documento
#print(dataframe)

data = []

# el rango empieza en 1 por que la primera fila corresponde al encabezado de el archivo en donde se indica
# el nombre de las columnas de cada grupo de datos
for row in range(1, dataframe.max_row):
    _row = [row,]
    # "row" nos indica el id de cada fila, o su número, es decir, hace el conteo de cuantas filas tenemos con información
    # print(row)
    for col in dataframe.iter_cols(1, dataframe.max_column):
        # imprime los datos fila por fila
        # print(col[row].value)
        _row.append(col[row].value)
    
    data.append(_row)

i = 0
cant = len(data[0])
for i in range(0, cant):
    
    print(data[0].pop(0))
    i += 1
    
    
